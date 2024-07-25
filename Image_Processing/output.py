import openpyxl
import os

from tqdm import tqdm

from Preprocess import preprocess
import pandas as pd


def output_xyl():
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    # pix unit = um
    result = preprocess(pix=0.098,
                        mask_path='K:/pythonProject/Torch_test/result/35.png',
                        image_path='K:/pythonProject/Torch_test/datas/test/predict/35.png',
                        file_id=0
                        )
    for temp in result[:len(result)]:
        for num_value in range(len(temp)):
            ws.cell(row=num_value+1, column=result.index(temp)+1, value=temp[num_value])

    # ws1 = wb.create_sheet('Gradient')
    # ws1.append(result[0])
    # for temp1 in result[-1]:
    #     for num_value1 in range(len(temp1)):
    #         ws1.cell(row=num_value1+2, column=result[-1].index(temp1)+2, value=temp1[num_value1])

    wb.save('output.csv')
    return None

def batch(mask_dir, image_dir, output_dir=None):
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    # 列出路径下的图片文件
    file_id = 0
    row_sum = 0
    mask_list = os.listdir(mask_dir)
    mask_list = [f for f in mask_list if any(ext in f.lower() for ext in ('.jpg', '.jpeg', '.png', '.bmp'))]
    # 进度条
    pbar = tqdm(total=len(mask_list), desc='Output')
    for img in mask_list:
        file_id += 1
        maskImg_dir = os.path.join(mask_dir, img)
        imageImg_dir = os.path.join(image_dir, img)
        # pix unit = um
        result = preprocess(pix=0.098,
                            mask_path=maskImg_dir,
                            image_path=imageImg_dir,
                            file_id=file_id
                            )

        for temp in result[:len(result)]:
            for num_value in range(0, len(temp)):
                ws.cell(row=num_value+row_sum+1, column=result.index(temp) + 1, value=temp[num_value])

        row_sum += len(result[0])
        pbar.update(1)

    wb.save(r'F:\Data\20240711\mengtuoshi\output.xlsx')
    return None

def idOrganize():
    # 读取excel文件
    df = pd.read_excel(r'F:\Data\20240711\mengtuoshi\output.xlsx')
    # 重新设置"ID"列作为正确的序号，从1开始
    df['number'] = range(1, len(df) + 1)
    # 将结果保存到新的Excel文件
    df.to_excel(r'F:\Data\20240711\mengtuoshi\output.xlsx', index=False)
    return None



if __name__ == '__main__':
    # input_dir = r'I:\0_Datas\Sediment'
    batch(r'F:\Data\20240711\mengtuoshi\binary', r'F:\Data\20240711\mengtuoshi\select')
    idOrganize()
