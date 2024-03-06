import openpyxl
from Preprocess import preprocess


def output_xyl():
    # 创建一个工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    pix = 0.098     # unit= um
    result = preprocess(pix)
    for temp in result[:len(result)]:
        for num_value in range(len(temp)):
            ws.cell(row=num_value+1, column=result.index(temp)+1, value=temp[num_value])

    # ws1 = wb.create_sheet('Gradient')
    # ws1.append(result[0])
    # for temp1 in result[-1]:
    #     for num_value1 in range(len(temp1)):
    #         ws1.cell(row=num_value1+2, column=result[-1].index(temp1)+2, value=temp1[num_value1])

    wb.save('D:\\Desktop\\test\\output\\output.xlsx')

    return None


if True:
    output_xyl()

